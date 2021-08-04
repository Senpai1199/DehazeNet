'''
    This script computes the NIQE (naturalness image quality evaluator) scores for the input images
    as well as the dehazed version of those images and saves the results in an excel file.
'''

from openpyxl import load_workbook
from util.get_path import get_path
import os
import cv2
from sewar.full_ref import uqi  # external lib to compute UQI score
from util.CLAHE import apply_CLAHE
from util.Adaptive_Enhance import adaptive_enhance
# from util.gamma_correction import gamma_correction
import numpy as np

with_coloration_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/testForImageCharacteristics/with-coloration/'
without_coloration_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/testForImageCharacteristics/without-coloration/'

dehazenet_coloration_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/images_with_coloration/'
dehazenet_without_coloration_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/images_without_coloration/'
gt_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/gt/'

excel_path = '/mnt/c/Users/Administrator/Desktop/criteria-test_Adaptive+CLAHE.xlsx'
wb = load_workbook(excel_path)
ws = wb['UQI']


def uqi_scores():
    i = 0
    for filename in os.listdir(with_coloration_path):
        i += 1
        gt_name = filename[:filename.find(
            '_')] + '.png'  # till the first occurence of underscore
        # print(gt_name)
        # print(gt_path + gt_name)
        gt_img = cv2.imread(gt_path + gt_name)
        im_path = with_coloration_path + filename
        inp_img = cv2.imread(im_path)
        predicted_path = get_path(inp_img, 117)  # threshold=117
        ws['A{}'.format(i + 1)] = filename
        ws['B{}'.format(i + 1)] = uqi(gt_img, inp_img)
        dehazenet_image = cv2.imread(
            dehazenet_coloration_path +
            '{}_finalWithoutCLAHE.jpg'.format(filename[:-4]))
        uqi_dehaze = uqi(gt_img, dehazenet_image)
        ws['C{}'.format(i + 1)] = uqi_dehaze

        if predicted_path == 1:
            ws['D{}'.format(i + 1)] = uqi(
                gt_img, apply_CLAHE(adaptive_enhance(inp_img)))
        else:
            ws['D{}'.format(i + 1)] = uqi_dehaze

        # Comparison with raw input image scores
        if ws['D{}'.format(i + 1)].value > ws['B{}'.format(i + 1)].value:
            ws['E{}'.format(i + 1)] = "Yes"
        elif ws['D{}'.format(i + 1)].value < ws['B{}'.format(i + 1)].value:
            ws['E{}'.format(i + 1)] = "No"
        else:
            ws['E{}'.format(i + 1)] = "Equal"

        # Comparison with DehazeNet scores
        if ws['D{}'.format(i + 1)].value > ws['C{}'.format(i + 1)].value:
            ws['F{}'.format(i + 1)] = "Yes"
        elif ws['D{}'.format(i + 1)].value < ws['C{}'.format(i + 1)].value:
            ws['F{}'.format(i + 1)] = "No"
        else:
            ws['F{}'.format(i + 1)] = "Equal"

        print("UQI computed for: ", filename)

    for filename in os.listdir(without_coloration_path):
        i += 1
        gt_name = filename[:filename.find(
            '_')] + '.png'  # till the first occurence of underscore
        gt_img = cv2.imread(gt_path + gt_name)
        im_path = without_coloration_path + filename
        inp_img = cv2.imread(im_path)
        predicted_path = get_path(inp_img, 117)  # threshold=117
        ws['A{}'.format(i + 1)] = filename
        ws['B{}'.format(i + 1)] = uqi(gt_img, inp_img)
        dehazenet_image = cv2.imread(
            dehazenet_without_coloration_path +
            '{}_finalWithoutCLAHE.jpg'.format(filename[:-4]))
        uqi_dehaze = uqi(gt_img, dehazenet_image)
        ws['C{}'.format(i + 1)] = uqi_dehaze

        if predicted_path == 1:
            ws['D{}'.format(i + 1)] = uqi(
                gt_img, apply_CLAHE(adaptive_enhance(inp_img)))
        else:
            ws['D{}'.format(i + 1)] = uqi_dehaze

        # Comparison with raw input image scores
        if ws['D{}'.format(i + 1)].value > ws['B{}'.format(i + 1)].value:
            ws['E{}'.format(i + 1)] = "Yes"
        elif ws['D{}'.format(i + 1)].value < ws['B{}'.format(i + 1)].value:
            ws['E{}'.format(i + 1)] = "No"
        else:
            ws['E{}'.format(i + 1)] = "Equal"

        # Comparison with DehazeNet scores
        if ws['D{}'.format(i + 1)].value > ws['C{}'.format(i + 1)].value:
            ws['F{}'.format(i + 1)] = "Yes"
        elif ws['D{}'.format(i + 1)].value < ws['C{}'.format(i + 1)].value:
            ws['F{}'.format(i + 1)] = "No"
        else:
            ws['F{}'.format(i + 1)] = "Equal"

        print("UQI computed for: ", filename)

    wb.save(excel_path)
    print("Script complete and excel saved!")


uqi_scores()
