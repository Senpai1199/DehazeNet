'''
    This script computes accuracies for different pixel threshold values for the get_path() function
    and stores the results in an excel file.
'''
from openpyxl import load_workbook
from criteria_test import compute_accuracy

excel_path = '/mnt/c/Users/Administrator/Desktop/criteria-test.xlsx'
wb = load_workbook(excel_path)
# print(wb.sheetnames)
ws = wb["Sheet2"]

i = 1
for threshold in range(110, 151):
    accuracy = compute_accuracy(threshold)
    print("Accuracy for {} is: {}".format(threshold, accuracy))
    ws['A{}'.format(i + 1)] = threshold
    ws['B{}'.format(i + 1)] = accuracy
    i += 1

wb.save(excel_path)