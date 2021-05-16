'''
    This script computes the accuracy for a particular threshold value for the pixel brightness
    for both with/without coloration input images to determine the path for dehazing.
    Also contains code for storing prediction correctness values in an excel sheet.
'''

from openpyxl import load_workbook
from util.get_path import get_path
import os
import cv2

with_coloration_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/testForImageCharacteristics/with-coloration/'
without_coloration_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/testForImageCharacteristics/without-coloration/'

excel_path = '/mnt/c/Users/Administrator/Desktop/criteria-test.xlsx'
wb = load_workbook(excel_path)
ws = wb['Sheet1']


def compute_accuracy(threshold=140):
    # print(wb.sheetnames)
    # print(ws['A1'].value)
    correct = 0
    i = 0
    for filename in os.listdir(with_coloration_path):
        print("Inside first for loop")
        i += 1
        im_path = with_coloration_path + filename
        predicted_path = get_path(cv2.imread(im_path), threshold)
        ws['C{}'.format(i + 1)] = predicted_path
        correct_path = ws['B{}'.format(i + 1)].value
        if predicted_path == correct_path:
            ws['D{}'.format(i + 1)] = "Yes"
            correct += 1
        else:
            # pass
            ws['D{}'.format(i + 1)] = "No"

    for filename in os.listdir(without_coloration_path):
        print("Inside second loop")
        i += 1
        im_path = without_coloration_path + filename
        predicted_path = get_path(
            cv2.imread(im_path),
            threshold)  # pass threshold as well to get_path()
        ws['C{}'.format(i + 1)] = predicted_path
        correct_path = ws['B{}'.format(i + 1)].value
        if predicted_path == correct_path:
            ws['D{}'.format(i + 1)] = "Yes"
            correct += 1
        else:
            # pass
            ws['D{}'.format(i + 1)] = "No"

    accuracy = (correct / i) * 100
    # return accuracy

    wb.save(excel_path)
    print("Script complete!")
    print("Path prediction Accuracy: ", accuracy)


compute_accuracy(140)
