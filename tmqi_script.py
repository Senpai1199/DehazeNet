'''
    This script computes the NIQE (naturalness image quality evaluator) scores for the input images
    as well as the dehazed version of those images and saves the results in an excel file.
'''

from openpyxl import load_workbook
from util.get_path import get_path
from util.tmqi import compute_tmqi
import os
import cv2
# from util.CLAHE import apply_CLAHE
from util.gamma_correction import gamma_correction
import numpy as np

with_coloration_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/testForImageCharacteristics/with-coloration/'
without_coloration_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/testForImageCharacteristics/without-coloration/'

dehazenet_coloration_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/images_with_coloration/'
dehazenet_without_coloration_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/images_without_coloration/'
gt_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/gt/'
enhanced_image_save_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/GammaCorrection1/'

excel_path = '/mnt/c/Users/Administrator/Desktop/criteria-test_GammaCorrection1.xlsx'
wb = load_workbook(excel_path)
ws = wb['TMQI-1']


def tmqi_scores():
    i = 0
    for filename in os.listdir(with_coloration_path):
        i += 1
        gt_name = filename[:filename.find(
            '_')] + '.png'  # till the first occurence of underscore
        # print(gt_name)
        # print(gt_path + gt_name)
        gt_img = gt_path + gt_name
        im_path = with_coloration_path + filename
        inp_img = cv2.imread(im_path)
        predicted_path = get_path(inp_img, 117)  # threshold=117
        ws['A{}'.format(i + 1)] = filename
        ws['B{}'.format(i + 1)] = compute_tmqi(gt_img, im_path)
        # dehazenet_image = dehazenet_coloration_path + '{}_finalWithoutCLAHE.jpg'.format(
        #     filename[:-4])
        # tmqi_dehaze = compute_tmqi(gt_img, dehazenet_image)
        # ws['C{}'.format(i + 1)] = tmqi_dehaze

        if predicted_path == 1:
            # equalized_img = dehazenet_coloration_path + '{}_Equalized.jpg'.format(
            #     filename[:-4])
            enh_path = enhanced_image_save_path + filename
            enhanced_img = cv2.imread(enhanced_image_save_path + filename)
            ws['D{}'.format(i + 1)] = compute_tmqi(gt_img, enh_path)
        else:
            ws['D{}'.format(i + 1)] = ws['C{}'.format(i + 1)].value

        # Comparison with raw input image scores
        if ws['D{}'.format(i + 1)].value > ws['B{}'.format(i + 1)].value:
            ws['E{}'.format(i + 1)] = "Yes"
        elif ws['D{}'.format(i + 1)].value < ws['B{}'.format(i + 1)].value:
            ws['E{}'.format(i + 1)] = "No"
        else:
            ws['E{}'.format(i + 1)] = "Equal"

        # Comparison with DehazeNet scores
        if ws['D{}'.format(i + 1)].value > ws['C{}'.format(i + 1)].value:
            ws['F{}'.format(i + 1)] = "Yes"
        elif ws['D{}'.format(i + 1)].value < ws['C{}'.format(i + 1)].value:
            ws['F{}'.format(i + 1)] = "No"
        else:
            ws['F{}'.format(i + 1)] = "Equal"

        print("TMQI computed for: ", filename)

    for filename in os.listdir(without_coloration_path):
        i += 1
        gt_name = filename[:filename.find(
            '_')] + '.png'  # till the first occurence of underscore
        gt_img = gt_path + gt_name
        im_path = without_coloration_path + filename
        inp_img = cv2.imread(im_path)
        predicted_path = get_path(inp_img, 117)  # threshold=117
        ws['A{}'.format(i + 1)] = filename
        ws['B{}'.format(i + 1)] = compute_tmqi(im_path, gt_img)
        # dehazenet_image = dehazenet_without_coloration_path + '{}_finalWithoutCLAHE.jpg'.format(
        #     filename[:-4])
        # tmqi_dehaze = compute_tmqi(dehazenet_image, gt_img)
        # ws['C{}'.format(i + 1)] = tmqi_dehaze

        if predicted_path == 1:
            # equalized_img = dehazenet_without_coloration_path + '{}_Equalized.jpg'.format(
            #     filename[:-4])
            enh_path = enhanced_image_save_path + filename
            enhanced_img = cv2.imread(enhanced_image_save_path + filename)
            ws['D{}'.format(i + 1)] = compute_tmqi(gt_img, enh_path)
        else:
            ws['D{}'.format(i + 1)] = ws['C{}'.format(i + 1)].value

        # Comparison with raw input image scores
        if ws['D{}'.format(i + 1)].value > ws['B{}'.format(i + 1)].value:
            ws['E{}'.format(i + 1)] = "Yes"
        elif ws['D{}'.format(i + 1)].value < ws['B{}'.format(i + 1)].value:
            ws['E{}'.format(i + 1)] = "No"
        else:
            ws['E{}'.format(i + 1)] = "Equal"

        # Comparison with DehazeNet scores
        if ws['D{}'.format(i + 1)].value > ws['C{}'.format(i + 1)].value:
            ws['F{}'.format(i + 1)] = "Yes"
        elif ws['D{}'.format(i + 1)].value < ws['C{}'.format(i + 1)].value:
            ws['F{}'.format(i + 1)] = "No"
        else:
            ws['F{}'.format(i + 1)] = "Equal"

        print("TMQI computed for: ", filename)

    wb.save(excel_path)
    print("Script complete and excel saved!")


tmqi_scores()
