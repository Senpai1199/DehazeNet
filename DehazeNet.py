import sys
import caffe
import numpy as np
import cv2
import math

import os
from skimage.metrics import structural_similarity
from openpyxl import load_workbook

from util.CLAHE import apply_CLAHE
from FADE import FADE as compute_fade


def EditFcnProto(templateFile, height, width):
    with open(templateFile, 'r') as ft:
        template = ft.read()
        outFile = 'DehazeNetFcn.prototxt'
        with open(outFile, 'w') as fd:
            fd.write(
                template.format(height_15=height + 15,
                                width_15=width + 15,
                                height_11=height + 11,
                                width_11=width + 11))


def TransmissionEstimate(im_path, height, width):
    caffe.set_mode_cpu()
    net = caffe.Net('DehazeNet.prototxt', 'DehazeNet.caffemodel', caffe.TEST)
    net_full_conv = caffe.Net('DehazeNetFcn.prototxt', 'DehazeNet.caffemodel',
                              caffe.TEST)
    net_full_conv.params['ip1-conv'][0].data.flat = net.params['ip1'][
        0].data.flat
    net_full_conv.params['ip1-conv'][1].data[...] = net.params['ip1'][1].data
    im = caffe.io.load_image(im_path)
    npad = ((7, 8), (7, 8), (0, 0))
    im = np.pad(im, npad, 'symmetric')
    transformers = caffe.io.Transformer(
        {'data': net_full_conv.blobs['data'].data.shape})
    transformers.set_transpose('data', (2, 0, 1))
    transformers.set_channel_swap('data', (2, 1, 0))
    out = net_full_conv.forward_all(
        data=np.array([transformers.preprocess('data', im - 0.2)]))
    transmission = np.reshape(out['ip1-conv'], (height, width))
    return transmission


def DarkChannel(im, sz):
    b, g, r = cv2.split(im)
    dc = cv2.min(cv2.min(r, g), b)
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (sz, sz))
    dark = cv2.erode(dc, kernel)
    return dark


def AtmLight(im, dark):
    [h, w] = im.shape[:2]
    imsz = h * w
    numpx = int(max(math.floor(imsz / 1000), 1))
    darkvec = dark.reshape(imsz, 1)
    imvec = im.reshape(imsz, 3)
    indices = darkvec.argsort()
    indices = indices[imsz - numpx::]
    atmsum = np.zeros([1, 3])
    for ind in range(1, numpx):
        atmsum = atmsum + imvec[indices[ind]]
    A = atmsum / numpx
    return A


def Guidedfilter(im, p, r, eps):
    mean_I = cv2.boxFilter(im, cv2.CV_64F, (r, r))
    mean_p = cv2.boxFilter(p, cv2.CV_64F, (r, r))
    mean_Ip = cv2.boxFilter(im * p, cv2.CV_64F, (r, r))
    cov_Ip = mean_Ip - mean_I * mean_p
    mean_II = cv2.boxFilter(im * im, cv2.CV_64F, (r, r))
    var_I = mean_II - mean_I * mean_I
    a = cov_Ip / (var_I + eps)
    b = mean_p - a * mean_I
    mean_a = cv2.boxFilter(a, cv2.CV_64F, (r, r))
    mean_b = cv2.boxFilter(b, cv2.CV_64F, (r, r))
    q = mean_a * im + mean_b
    return q


def TransmissionRefine(im, et):
    gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
    gray = np.float64(gray) / 255
    r = 60
    eps = 0.0001
    t = Guidedfilter(gray, et, r, eps)
    return t


def Recover(im, t, A, tx=0.1):
    res = np.empty(im.shape, im.dtype)
    t = cv2.max(t, tx)
    for ind in range(0, 3):
        res[:, :, ind] = (im[:, :, ind] - A[0, ind]) / t + A[0, ind]
    return res


# if __name__ == '__main__':
#     if not len(sys.argv) == 2:
#         print('Usage: python DehazeNet.py haze_img_path')
#         exit()
#     else:
#         im_path = sys.argv[1]
#     src = cv2.imread(im_path)
#     # src = apply_CLAHE(src)  # applying CLAHE
#     height = src.shape[0]
#     width = src.shape[1]

#     equalized_img_path = im_path[:-4] + '_Equalized' + im_path[-4:len(im_path)]
#     cv2.imwrite(equalized_img_path, src * 255)
#     print("Equalized image saved")

#     templateFile = 'DehazeFcnTemplate.prototxt'
#     EditFcnProto(templateFile, height, width)
#     I = src / 255.0
#     dark = DarkChannel(I, 15)
#     A = AtmLight(I, dark)
#     te = TransmissionEstimate(im_path, height, width)
#     t = TransmissionRefine(src, te)
#     J = Recover(I, t, A, 0.1)
#     print("HEREE")
#     # cv2.imshow('TransmissionEstimate',te)
#     # cv2.imshow('TransmissionRefine',t)
#     # cv2.imshow('Origin',src)
#     # cv2.imshow('Dehaze',J)
#     cv2.waitKey(0)
#     save_path = im_path[:-4] + '_Dehaze' + im_path[-4:len(im_path)]
#     cv2.imwrite(save_path, J * 255)

#     print("Image saved at: ", save_path)


def dehaze_image(src, im_path):
    """
        Receives a hazy image and returns the dehazed image after passing through DehazeNet
    """
    height = src.shape[0]
    width = src.shape[1]

    # equalized_img_path = im_path[:-4] + '_Equalized' + im_path[-4:len(im_path)]
    # cv2.imwrite(equalized_img_path, src * 255)
    # print("Equalized image saved")

    templateFile = 'DehazeFcnTemplate.prototxt'
    EditFcnProto(templateFile, height, width)
    I = src / 255.0
    dark = DarkChannel(I, 15)
    A = AtmLight(I, dark)
    te = TransmissionEstimate(im_path, height, width)
    t = TransmissionRefine(src, te)
    J = Recover(I, t, A, 0.1)
    # cv2.waitKey(0)
    # save_path = im_path[:-4] + '_Dehaze' + im_path[-4:len(im_path)]
    final_img = J * 255
    # cv2.imwrite(save_path, final_img)
    return final_img


base_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/'
hazy_path = base_path + 'hazy/'
# results_path_withCLAHE = base_path + 'results-withCLAHE/'
# results_path_withoutCLAHE = base_path + 'results-withoutCLAHE/'

# Open excel file
excel_path = '/mnt/c/Users/Administrator/Desktop/test.xlsx'
wb = load_workbook(excel_path)
ws = wb.active

# driver code to generate metrics for all 500 images

# 1. For FADE scores
# i = 0
# for filename in os.listdir(hazy_path):
#     i += 1
#     if (i <= 2):
#         im_path = hazy_path + filename
#         src = cv2.imread(im_path)
#         fade_score = compute_fade(src)
#         # src = cv2.imread(os.path.join(results_path_withCLAHE, filename))
#         # src_with_CLAHE = apply_CLAHE(src)
#         # # save_path = base_path + 'results/' + filename[:
#         #   -4] + '_Equalized' + filename[
#         #       -4:len(filename)]
#         # cv2.imwrite(save_path, src_with_CLAHE)
#         # dehazed_without_CLAHE = dehaze_image(src, im_path)
#         # dehazed_with_CLAHE = dehaze_image(src_with_CLAHE, im_path)
#         # save_path = base_path + 'results/' + filename[:
#         #   -4] + '_finalWithCLAHE' + filename[
#         #       -4:len(filename)]
#         # cv2.imwrite(save_path, dehazed_with_CLAHE)

#         # save_path = base_path + 'results/' + filename[:
#         #   -4] + '_finalWithoutCLAHE' + filename[
#         #       -4:len(filename)]
#         # cv2.imwrite(save_path, dehazed_without_CLAHE)

#         # ssim_without_CLAHE = structural_similarity(src,
#         #                                            dehazed_without_CLAHE,
#         #                                            multichannel=True)
#         # ssim_with_CLAHE = structural_similarity(src_with_CLAHE,
#         # dehazed_with_CLAHE,
#         # multichannel=True)
#         ws['A{}'.format(i + 1)] = filename
#         ws['B{}'.format(i + 1)] = fade_score
#         # ws['E{}'.format(i + 1)] = ssim_with_CLAHE
#         print("Image No.: ", i)
#         print("FILENAME: ", filename)
#         print("FADE SCORE WITH CLAHE: ", fade_score)
#         # print("SSIM WITH CLAHE: ", ssim_with_CLAHE)
# wb.save('test.xlsx')

# i = 0
# for filename in os.listdir(results_path_withoutCLAHE):
#     i += 1
#     if (i <= 2):
#         im_path = results_path_withoutCLAHE + filename
#         src = cv2.imread(im_path)
#         fade_score = compute_fade(src)
#         # src = cv2.imread(os.path.join(results_path_withoutCLAHE, filename))
#         # src_with_CLAHE = apply_CLAHE(src)
#         # # save_path = base_path + 'results/' + filename[:
#         #   -4] + '_Equalized' + filename[
#         #       -4:len(filename)]
#         # cv2.imwrite(save_path, src_with_CLAHE)
#         # dehazed_without_CLAHE = dehaze_image(src, im_path)
#         # dehazed_with_CLAHE = dehaze_image(src_with_CLAHE, im_path)
#         # save_path = base_path + 'results/' + filename[:
#         #   -4] + '_finalWithCLAHE' + filename[
#         #       -4:len(filename)]
#         # cv2.imwrite(save_path, dehazed_with_CLAHE)

#         # save_path = base_path + 'results/' + filename[:
#         #   -4] + '_finalWithoutCLAHE' + filename[
#         #       -4:len(filename)]
#         # cv2.imwrite(save_path, dehazed_without_CLAHE)

#         # ssim_without_CLAHE = structural_similarity(src,
#         #                                            dehazed_without_CLAHE,
#         #                                            multichannel=True)
#         # ssim_with_CLAHE = structural_similarity(src_with_CLAHE,
#         # dehazed_with_CLAHE,
#         # multichannel=True)
#         ws['C{}'.format(i + 1)] = fade_score
#         # ws['E{}'.format(i + 1)] = ssim_with_CLAHE
#         print("Image No.: ", i)
#         print("FILENAME: ", filename)
#         print("FADE SCORE WITHOUT CLAHE: ", fade_score)
#         # print("SSIM WITH CLAHE: ", ssim_with_CLAHE)
# wb.save('test.xlsx')

# 2. For SSIM and FADE
i = 0
for filename in os.listdir(hazy_path):
    i += 1
    # if (i <= 2):
    im_path = hazy_path + filename
    src = cv2.imread(im_path)
    src_with_CLAHE = apply_CLAHE(src)

    save_path = base_path + 'results_test1/' + filename[:-4] + filename[
        -4:len(filename)]
    cv2.imwrite(
        save_path,
        src)  # saving original image in the results_test1 folder as well

    save_path = base_path + 'results_test1/' + filename[:
                                                        -4] + '_Equalized' + filename[
                                                            -4:len(filename)]
    cv2.imwrite(save_path, src_with_CLAHE)
    dehazed_without_CLAHE = dehaze_image(src, im_path)
    dehazed_with_CLAHE = dehaze_image(src_with_CLAHE, im_path)
    save_path = base_path + 'results_test1/' + filename[:
                                                        -4] + '_finalWithCLAHE' + filename[
                                                            -4:len(filename)]
    cv2.imwrite(save_path, dehazed_with_CLAHE)

    save_path = base_path + 'results_test1/' + filename[:
                                                        -4] + '_finalWithoutCLAHE' + filename[
                                                            -4:len(filename)]
    cv2.imwrite(save_path, dehazed_without_CLAHE)

    ssim_without_CLAHE = structural_similarity(src,
                                               dehazed_without_CLAHE,
                                               multichannel=True)
    ssim_with_CLAHE = structural_similarity(src_with_CLAHE,
                                            dehazed_with_CLAHE,
                                            multichannel=True)
    fade_score_withoutCLAHE = compute_fade(dehazed_without_CLAHE)
    fade_score_withCLAHE = compute_fade(dehazed_with_CLAHE)
    # fade_score = compute_fade(src)
    ws['A{}'.format(i + 1)] = filename
    # ws['B{}'.format(i + 1)] = fade_score
    print("Image no: ", i)
    ws['B{}'.format(i + 1)] = fade_score_withCLAHE
    ws['C{}'.format(i + 1)] = fade_score_withoutCLAHE
    ws['D{}'.format(i + 1)] = ssim_with_CLAHE
    ws['E{}'.format(i + 1)] = ssim_without_CLAHE
    # ws['D{}'.format(i + 1)] = fade_score
    # print("Image No.: ", i)
    # print("FILENAME: ", filename)
    # print("FADE SCORE WITH CLAHE: ", fade_score)
    # print("SSIM WITH CLAHE: ", ssim_with_CLAHE)
wb.save(excel_path)
print("Excel saved!")
