'''
    Script to compute Red, Green, Blue channel metrics such as standard deviation
    and mean pixel brightness values. Also contains code for plotting and storing histograms for 
    the three channel pixel brighntess values separately for each input image.
'''
# import cv2
import os
import numpy as np
from openpyxl import load_workbook
from PIL import Image
import matplotlib.pyplot as plt

with_coloration_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/testForImageCharacteristics/with-coloration/'
without_coloration_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/testForImageCharacteristics/without-coloration/'
# hist_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/testForImageCharacteristics/histograms/'

excel_path = '/mnt/c/Users/Administrator/Desktop/image-char_test.xlsx'
wb = load_workbook(excel_path)
ws = wb.active

i = 0
# with coloration loop
for filename in os.listdir(with_coloration_path):
    i += 1
    im_path = with_coloration_path + filename
    im_array = np.array(Image.open(im_path))
    red_chan = np.ndarray.flatten(im_array[:, :, 0])
    green_chan = np.ndarray.flatten(im_array[:, :, 1])
    blue_chan = np.ndarray.flatten(im_array[:, :, 2])

    # red_hist = plt.hist(red_chan)
    # plt.xlabel('Brightness value (0->Dark, 255->White)')
    # plt.ylabel('Frequency')
    # plt.savefig(hist_path + "with-coloration/red/" + filename + "_red.png")
    # plt.clf()
    # green_hist = plt.hist(green_chan)
    # plt.xlabel('Brightness value (0->Dark, 255->White)')
    # plt.ylabel('Frequency')
    # plt.savefig(hist_path + "with-coloration/green/" + filename + "_green.png")
    # plt.clf()
    # blue_hist = plt.hist(blue_chan)
    # plt.xlabel('Brightness value (0->Dark, 255->White)')
    # plt.ylabel('Frequency')
    # plt.savefig(hist_path + "with-coloration/blue/" + filename + "_blue.png")
    # plt.clf()

    mean_red = np.mean(red_chan)
    mean_green = np.mean(green_chan)
    mean_blue = np.mean(blue_chan)

    stddev_red = np.std(red_chan)
    stddev_green = np.std(green_chan)
    stddev_blue = np.std(blue_chan)

    ws['A{}'.format(i + 1)] = filename
    ws['B{}'.format(i + 1)] = 'with'
    ws['C{}'.format(i + 1)] = stddev_red
    ws['D{}'.format(i + 1)] = stddev_green
    ws['E{}'.format(i + 1)] = stddev_blue
    ws['F{}'.format(i + 1)] = mean_red
    ws['G{}'.format(i + 1)] = mean_green
    ws['H{}'.format(i + 1)] = mean_blue

# without coloration loop
for filename in os.listdir(without_coloration_path):
    i += 1
    im_path = without_coloration_path + filename
    im_array = np.array(Image.open(im_path))
    red_chan = np.ndarray.flatten(im_array[:, :, 0])
    green_chan = np.ndarray.flatten(im_array[:, :, 1])
    blue_chan = np.ndarray.flatten(im_array[:, :, 2])

    # red_hist = plt.hist(red_chan)
    # plt.xlabel('Brightness value (0->Dark, 255->White)')
    # plt.ylabel('Frequency')
    # plt.savefig(hist_path + "without-coloration/red/" + filename + "_red.png")
    # plt.clf()
    # green_hist = plt.hist(green_chan)
    # plt.xlabel('Brightness value (0->Dark, 255->White)')
    # plt.ylabel('Frequency')
    # plt.savefig(hist_path + "without-coloration/green/" + filename +
    #             "_green.png")
    # plt.clf()
    # blue_hist = plt.hist(blue_chan)
    # plt.xlabel('Brightness value (0->Dark, 255->White)')
    # plt.ylabel('Frequency')
    # plt.savefig(hist_path + "without-coloration/blue/" + filename +
    #             "_blue.png")
    # plt.clf()

    mean_red = np.mean(red_chan)
    mean_green = np.mean(green_chan)
    mean_blue = np.mean(blue_chan)

    stddev_red = np.std(red_chan)
    stddev_green = np.std(green_chan)
    stddev_blue = np.std(blue_chan)
    ws['A{}'.format(i + 1)] = filename
    ws['B{}'.format(i + 1)] = 'without'
    ws['C{}'.format(i + 1)] = stddev_red
    ws['D{}'.format(i + 1)] = stddev_green
    ws['E{}'.format(i + 1)] = stddev_blue
    ws['F{}'.format(i + 1)] = mean_red
    ws['G{}'.format(i + 1)] = mean_green
    ws['H{}'.format(i + 1)] = mean_blue

wb.save(excel_path)
print("Script complete!")
