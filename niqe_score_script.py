'''
    This script computes the NIQE (naturalness image quality evaluator) scores for the input images
    as well as the dehazed version of those images and saves the results in an excel file.
'''

from openpyxl import load_workbook
from util.get_path import get_path
import os
import cv2
from util.niqe import niqe
from util.CLAHE import apply_CLAHE
# from util.gamma_correction import gamma_correction
from util.Adaptive_Enhance import adaptive_enhance
import numpy as np

with_coloration_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/testForImageCharacteristics/with-coloration/'
without_coloration_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/testForImageCharacteristics/without-coloration/'

dehazenet_coloration_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/images_with_coloration/'
dehazenet_without_coloration_path = '/mnt/c/Users/Administrator/Desktop/SOTS/outdoor/images_without_coloration/'

excel_path = '/mnt/c/Users/Administrator/Desktop/criteria-test_Adaptive+CLAHE.xlsx'
wb = load_workbook(excel_path)
ws = wb['NIQE']


def niqe_scores():
    i = 0
    for filename in os.listdir(with_coloration_path):
        # print("Filename: ", filename)
        # print("Inside first for loop")
        i += 1
        im_path = with_coloration_path + filename
        inp_img = cv2.imread(im_path)
        niqe_inp_img = niqe(inp_img)
        predicted_path = get_path(inp_img, 117)  # threshold=117
        ws['A{}'.format(i + 1)] = filename
        ws['B{}'.format(i + 1)] = niqe_inp_img
        dehazenet_image = cv2.imread(
            dehazenet_coloration_path +
            '{}_finalWithoutCLAHE.jpg'.format(filename[:-4]))
        niqe_dehazenet = niqe(dehazenet_image)
        ws['C{}'.format(i + 1)] = niqe_dehazenet

        if predicted_path == 1:
            enhanced_img = apply_CLAHE(adaptive_enhance(inp_img))
            ws['D{}'.format(i + 1)] = niqe(enhanced_img)
        else:
            ws['D{}'.format(i + 1)] = niqe_dehazenet

        # Comparison with raw input image scores
        # In NIQE, lesser score is better
        if ws['D{}'.format(i + 1)].value < ws['B{}'.format(i + 1)].value:
            ws['E{}'.format(i + 1)] = "Yes"
        elif ws['D{}'.format(i + 1)].value > ws['B{}'.format(i + 1)].value:
            ws['E{}'.format(i + 1)] = "No"
        else:
            ws['E{}'.format(i + 1)] = "Equal"

        # Comparison with DehazeNet scores
        if ws['D{}'.format(i + 1)].value < ws['C{}'.format(i + 1)].value:
            ws['F{}'.format(i + 1)] = "Yes"
        elif ws['D{}'.format(i + 1)].value > ws['C{}'.format(i + 1)].value:
            ws['F{}'.format(i + 1)] = "No"
        else:
            ws['F{}'.format(i + 1)] = "Equal"
            # ws['D{}'.format(i + 1)] = niqe(dehaze_image(inp_img, im_path))
            # ws['C{}'.format(i + 1)] = ws['D{}'.format(i + 1)].value

        print("NIQE computed for: ", filename)

    for filename in os.listdir(without_coloration_path):
        # print("Filename: ", filename)
        # print("Inside first for loop")
        i += 1
        im_path = without_coloration_path + filename
        inp_img = cv2.imread(im_path)
        niqe_inp_img = niqe(inp_img)
        predicted_path = get_path(inp_img, 117)  # threshold=117
        ws['A{}'.format(i + 1)] = filename
        ws['B{}'.format(i + 1)] = niqe_inp_img
        dehazenet_image = cv2.imread(
            dehazenet_without_coloration_path +
            '{}_finalWithoutCLAHE.jpg'.format(filename[:-4]))
        niqe_dehazenet = niqe(dehazenet_image)
        ws['C{}'.format(i + 1)] = niqe_dehazenet

        if predicted_path == 1:
            enhanced_img = apply_CLAHE(adaptive_enhance(inp_img))
            ws['D{}'.format(i + 1)] = niqe(enhanced_img)
        else:
            ws['D{}'.format(i + 1)] = niqe_dehazenet

        # Comparison with raw input image scores
        # In NIQE, lesser score is better
        if ws['D{}'.format(i + 1)].value < ws['B{}'.format(i + 1)].value:
            ws['E{}'.format(i + 1)] = "Yes"
        elif ws['D{}'.format(i + 1)].value > ws['B{}'.format(i + 1)].value:
            ws['E{}'.format(i + 1)] = "No"
        else:
            ws['E{}'.format(i + 1)] = "Equal"

        # Comparison with DehazeNet scores
        if ws['D{}'.format(i + 1)].value < ws['C{}'.format(i + 1)].value:
            ws['F{}'.format(i + 1)] = "Yes"
        elif ws['D{}'.format(i + 1)].value > ws['C{}'.format(i + 1)].value:
            ws['F{}'.format(i + 1)] = "No"
        else:
            ws['F{}'.format(i + 1)] = "Equal"

            # op_img = dehaze_image(inp_img, im_path)
            # niqe_score = niqe(op_img)
            # ws['D{}'.format(i + 1)] = niqe_score
            # ws['C{}'.format(i + 1)] = ws['D{}'.format(i + 1)].value
        print("NIQE computed for: ", filename)

    wb.save(excel_path)
    print("Script complete and excel saved!")


niqe_scores()
